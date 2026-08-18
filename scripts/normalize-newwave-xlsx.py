from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


def text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def cents(value):
    value = text(value)
    if not value:
        return None
    try:
        return int(round(float(value.replace(",", ".")) * 100))
    except ValueError:
        return None


def split_list(value):
    value = text(value)
    if not value:
        return []
    return [part.strip() for part in re.split(r"[,;|]", value) if part.strip()]


def document_type(url):
    normalized = url.lower()
    if any(token in normalized for token in ["storlek", "size", "talla"]):
        return "SIZE_GUIDE"
    if any(token in normalized for token in ["declaration", "conformity", "cert", "doc"]):
        return "TECHNICAL_SHEET"
    return "DOCUMENT"


def document_title(doc_type):
    return {
        "SIZE_GUIDE": "Guía de tallas",
        "TECHNICAL_SHEET": "Ficha técnica",
        "DOCUMENT": "Documento técnico",
    }[doc_type]


def slugify(value):
    value = unicodedata.normalize("NFD", value).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[^a-zA-Z0-9]+", "-", value.lower()).strip("-")
    return value or "producte"


def first_value(row, indexes, names):
    for name in names:
        index = indexes.get(name)
        if index is not None and index < len(row):
            value = text(row[index])
            if value:
                return value
    return None


def spec(label, value):
    value = text(value)
    if not value:
        return None
    return {"label": label, "value": value}


def normalize(xlsx_path: Path, output_path: Path):
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    products = OrderedDict()
    skipped = []

    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        indexes = {str(value).strip(): index for index, value in enumerate(header) if text(value)}

        for row_number, row in enumerate(rows, start=2):
            if not any(text(value) for value in row):
                continue

            product_number = first_value(row, indexes, ["Product number"])
            sku = first_value(row, indexes, ["Sku"])
            name = first_value(row, indexes, ["Product name (es)"])
            brand = first_value(row, indexes, ["Brand"]) or worksheet.title
            price = cents(first_value(row, indexes, ["Precio recomendado"]))

            if not product_number or not sku or not name or not price or price <= 0:
                skipped.append(
                    {
                        "sheet": worksheet.title,
                        "row": row_number,
                        "productNumber": product_number,
                        "sku": sku,
                        "name": name,
                        "price": price,
                    }
                )
                continue

            categories = split_list(first_value(row, indexes, ["Category (es)"]))
            description = first_value(row, indexes, ["Description, Catalog (es)", "Description (es)", "Description, Retail (es)"])
            image_sources = []
            for column in ["Main image", "Product images", "All images"]:
                image_sources.extend(split_list(first_value(row, indexes, [column])))
            images = []
            for url in image_sources:
                if url.startswith("http") and url not in images:
                    images.append(url)
            documents = []
            for url in split_list(first_value(row, indexes, ["Documents"])):
                if not url.startswith("http") or any(document["url"] == url for document in documents):
                    continue
                doc_type = document_type(url)
                documents.append({"type": doc_type, "title": document_title(doc_type), "url": url})

            specifications = [
                item
                for item in [
                    spec("Género", first_value(row, indexes, ["Gender (es)"])),
                    spec("Composición", first_value(row, indexes, ["Fabrics (es)"])),
                    spec("Gramaje", first_value(row, indexes, ["Textile weight"])),
                    spec("Técnica/material", first_value(row, indexes, ["Material technique (es)"])),
                    spec("Certificación", first_value(row, indexes, ["Certification (es)"])),
                    spec("Descripción de certificación", first_value(row, indexes, ["Certification description (es)"])),
                    spec("Uso recomendado", first_value(row, indexes, ["Activity type (es)"])),
                    spec("Área de aplicación", first_value(row, indexes, ["Application area (es)"])),
                    spec("Tipo de aplicación", first_value(row, indexes, ["Application type (es)"])),
                    spec("Información de aplicación", first_value(row, indexes, ["Application info (es)"])),
                    spec("Características", first_value(row, indexes, ["Feature (es)"])),
                    spec("Ajuste", first_value(row, indexes, ["Fit (es)"])),
                    spec("Cierre", first_value(row, indexes, ["Closure (es)"])),
                    spec("Bolsillos", first_value(row, indexes, ["Pockets (es)"])),
                    spec("Capucha", first_value(row, indexes, ["Hood details (es)"])),
                    spec("Cuello", first_value(row, indexes, ["Neck line (es)"])),
                    spec("Manga", first_value(row, indexes, ["Sleeve (es)"])),
                    spec("Instrucciones de cuidado", first_value(row, indexes, ["Care instructions (es)"])),
                    spec("Comentario técnico", first_value(row, indexes, ["Technique comment (es)"])),
                    spec("Comentario de color", first_value(row, indexes, ["Color comment (es)"])),
                    spec("Packaging", first_value(row, indexes, ["Packaging Comment (es)"])),
                    spec("Peso", first_value(row, indexes, ["Weight"])),
                    spec("Código de impresión", first_value(row, indexes, ["Print code"])),
                    spec("Temporada", first_value(row, indexes, ["Season"])),
                    spec("Rango", first_value(row, indexes, ["Range (es)"])),
                    spec("USP", first_value(row, indexes, ["USP text (es)"])),
                ]
                if item
            ]

            if product_number not in products:
                gender = first_value(row, indexes, ["Gender (es)"])
                material = first_value(row, indexes, ["Material technique (es)"])
                products[product_number] = {
                    "productNumber": product_number,
                    "name": name,
                    "slug": f"{slugify(name)}-{slugify(product_number)}",
                    "description": description,
                    "gender": gender,
                    "material": material,
                    "category": categories[0] if categories else None,
                    "subcategory": categories[1] if len(categories) > 1 else None,
                    "brand": brand,
                    "images": images[:12],
                    "documents": documents,
                    "specifications": specifications,
                    "attributes": {
                        key: value
                        for key, value in {
                            "gender": first_value(row, indexes, ["Gender (es)"]),
                            "fabrics": first_value(row, indexes, ["Fabrics (es)"]),
                            "certification": first_value(row, indexes, ["Certification (es)"]),
                            "activityType": first_value(row, indexes, ["Activity type (es)"]),
                            "applicationType": first_value(row, indexes, ["Application type (es)"]),
                        }.items()
                        if value
                    },
                    "variants": [],
                }
            else:
                existing_images = products[product_number]["images"]
                for url in images:
                    if url not in existing_images and len(existing_images) < 12:
                        existing_images.append(url)
                existing_documents = products[product_number]["documents"]
                for document in documents:
                    if all(existing["url"] != document["url"] for existing in existing_documents):
                        existing_documents.append(document)

            color = first_value(row, indexes, ["Color (es)", "Color code"])
            size = first_value(row, indexes, ["Size name", "Size code"])
            products[product_number]["variants"].append(
                {
                    "sku": sku,
                    "name": " / ".join(part for part in [color, size] if part) or sku,
                    "color": color,
                    "size": size,
                    "priceCents": price,
                    "stock": 999,
                }
            )

    normalized = []
    for product in products.values():
        seen_skus = set()
        variants = []
        for variant in product["variants"]:
            if variant["sku"] in seen_skus:
                continue
            seen_skus.add(variant["sku"])
            variants.append(variant)
        product["variants"] = variants
        normalized.append(product)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(normalized, ensure_ascii=False), encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(output_path),
                "products": len(normalized),
                "variants": sum(len(product["variants"]) for product in normalized),
                "skippedRows": len(skipped),
                "skippedExamples": skipped[:10],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python scripts/normalize-newwave-xlsx.py <input.xlsx> <output.json>")
    normalize(Path(sys.argv[1]), Path(sys.argv[2]))
