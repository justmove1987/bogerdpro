from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


def text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def cents(value):
    value = text(value)
    if not value:
        return None
    try:
        return int(round(float(value.replace(",", ".")) * 100))
    except ValueError:
        return None


def split_urls(value):
    value = text(value)
    if not value:
        return []
    urls = []
    for part in re.split(r"[,;|]", value):
        part = part.strip()
        if part.startswith("http") and part not in urls:
            urls.append(part)
    return urls


def slugify(value):
    value = unicodedata.normalize("NFD", value).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[^a-zA-Z0-9]+", "-", value.lower()).strip("-")
    return value or "producte"


def clean_color(value):
    value = text(value)
    return " ".join(value.split()).title() if value else None


def first_sentence_name(name):
    value = text(name)
    if not value:
        return None
    return " ".join(value.split())


def spec(label, value):
    value = text(value)
    if not value:
        return None
    return {"label": label, "value": value}


def normalize(xlsx_path: Path, output_path: Path):
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    next(rows, None)
    header = [str(value).strip() if value is not None else "" for value in next(rows)]
    indexes = {name: index for index, name in enumerate(header) if name}
    products = OrderedDict()
    skipped = []

    for row_number, row in enumerate(rows, start=3):
        raw = {name: row[index] if index < len(row) else None for name, index in indexes.items()}
        product_url = text(row[23]) if len(row) > 23 else None
        if not any(text(value) for value in raw.values()):
            continue

        model = text(raw.get("Producto/Modelo"))
        variant_sku = text(raw.get("Referencia interna"))
        barcode = text(raw.get("Código de barras")) or text(raw.get("C�digo de barras"))
        name = first_sentence_name(raw.get("Nombre comercial"))
        price = cents(raw.get("Precio público de Página Web")) or cents(raw.get("Precio p�blico de P�gina Web"))
        color = clean_color(raw.get("Color"))
        size = text(raw.get("Talla"))

        if not model or not variant_sku or not name or not price or price <= 0:
            skipped.append({"row": row_number, "model": model, "sku": variant_sku, "name": name, "price": price})
            continue

        category = text(raw.get("Por colección")) or text(raw.get("Por colecci�n"))
        subcategory = text(raw.get("Por artículo")) or text(raw.get("Por art�culo"))
        description = text(raw.get("Descripción corta")) or text(raw.get("Descripci�n corta")) or text(raw.get("Descripción para venta")) or text(raw.get("Descripci�n para venta"))
        short_description = text(raw.get("Descripción para venta")) or text(raw.get("Descripci�n para venta"))
        composition = text(raw.get("Composición")) or text(raw.get("Composici�n"))
        treatment = text(raw.get("Tratamiento"))
        profession = text(raw.get("Por profesión")) or text(raw.get("Por profesi�n"))
        collection = text(raw.get("Por colección")) or text(raw.get("Por colecci�n"))
        images = []
        for url in split_urls(raw.get("URL imagen")) + split_urls(raw.get("Enlace descarga fotos")):
            if url not in images:
                images.append(url)

        if model not in products:
            documents = []
            if product_url and product_url.startswith("http"):
                documents.append({"type": "DOCUMENT", "title": "Ficha de producto", "url": product_url})

            specifications = [
                item
                for item in [
                    spec("Composición", composition),
                    spec("Tratamiento", treatment),
                    spec("Profesión", profession),
                    spec("Colección", collection),
                    spec("Artículo", subcategory),
                    spec("Peso", raw.get("Peso")),
                    spec("Volumen", raw.get("Volumen")),
                    spec("Código arancelario", raw.get("Código arancelario") or raw.get("C�digo arancelario")),
                    spec("Nombre arancelario", raw.get("Nombre arancelario")),
                    spec("Vídeo", raw.get("URL vídeo") or raw.get("URL v�deo")),
                ]
                if item
            ]

            products[model] = {
                "productNumber": model,
                "name": name,
                "slug": f"{slugify(name)}-{slugify(model)}",
                "description": description or short_description,
                "gender": None,
                "material": composition,
                "category": category,
                "subcategory": subcategory,
                "brand": "Gary's",
                "images": images[:12],
                "documents": documents,
                "specifications": specifications,
                "attributes": {
                    key: value
                    for key, value in {
                        "profession": profession,
                        "article": subcategory,
                        "collection": collection,
                        "composition": composition,
                        "treatment": treatment,
                    }.items()
                    if value
                },
                "variants": [],
            }
        else:
            existing_images = products[model]["images"]
            for url in images:
                if url not in existing_images and len(existing_images) < 12:
                    existing_images.append(url)

        products[model]["variants"].append(
            {
                "sku": variant_sku,
                "name": " / ".join(part for part in [color, size] if part) or variant_sku,
                "color": color,
                "size": size,
                "priceCents": price,
                "stock": 999,
                "barcode": barcode,
            }
        )

    normalized = []
    for product in products.values():
        seen_skus = set()
        variants = []
        for variant in product["variants"]:
            if variant["sku"] in seen_skus:
                continue
            seen_skus.add(variant["sku"])
            variant.pop("barcode", None)
            variants.append(variant)
        product["variants"] = variants
        if variants:
            normalized.append(product)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Normalized {len(normalized)} products with {sum(len(item['variants']) for item in normalized)} variants.")
    if skipped:
        print(f"Skipped {len(skipped)} rows without model, SKU, name or price.")
        print(json.dumps(skipped[:10], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python scripts/normalize-garys-xlsx.py <input.xlsx> <output.json>")
    normalize(Path(sys.argv[1]), Path(sys.argv[2]))
