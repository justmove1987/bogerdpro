from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


def text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def cents(value):
    value = text(value)
    if not value:
        return None
    try:
        return int(round(float(value.replace(",", ".")) * 100))
    except ValueError:
        return None


def slugify(value):
    value = unicodedata.normalize("NFD", value).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[^a-zA-Z0-9]+", "-", value.lower()).strip("-")
    return value or "producte"


def clean_title(value):
    value = text(value)
    if not value:
        return None
    return " ".join(value.split())


def clean_color(value):
    value = clean_title(value)
    return value.title() if value else None


def clean_size(value):
    value = clean_title(value)
    if not value:
        return None
    replacements = {
        "TALLA �NICA ADULTO": "Talla única adulto",
        "TALLA ÚNICA ADULTO": "Talla única adulto",
        "TALLA �NICA INFANTIL": "Talla única infantil",
        "TALLA ÚNICA INFANTIL": "Talla única infantil",
        "UNICA": "Única",
    }
    return replacements.get(value.upper(), value)


def category_from_categories(value):
    parts = [part.strip() for part in (text(value) or "").split(",") if part.strip()]
    ignored = {"stamina", "merchandising", "novedades", "verano", "sublimaci�n", "sublimación"}
    for part in parts:
        normalized = unicodedata.normalize("NFD", part).encode("ascii", "ignore").decode("ascii").lower()
        if normalized not in ignored:
            return part.title()
    return parts[0].title() if parts else None


def spec(label, value):
    value = text(value)
    if not value:
        return None
    return {"label": label, "value": value}


def normalize(xlsx_path: Path, output_path: Path):
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = [str(value).strip() if value is not None else "" for value in next(rows)]
    indexes = {name: index for index, name in enumerate(header) if name}
    products = OrderedDict()
    skipped = []

    for row_number, row in enumerate(rows, start=2):
        raw = {name: row[index] if index < len(row) else None for name, index in indexes.items()}
        if not any(text(value) for value in raw.values()):
            continue

        sku = text(raw.get("PRODUCTCODE"))
        model = text(raw.get("MODELCODE"))
        name = clean_title(raw.get("MODELNAME"))
        price = cents(raw.get("PRICE < 500 UNITS"))
        color = clean_color(raw.get("COLOR"))
        size = clean_size(raw.get("SIZE"))

        if not sku or not model or not name or not price or price <= 0:
            skipped.append({"row": row_number, "model": model, "sku": sku, "name": name, "price": price})
            continue

        product_image = text(raw.get("PRODUCTIMAGE"))
        model_image = text(raw.get("MODELIMAGE"))
        images = []
        for url in [product_image, model_image]:
            if url and url.startswith("http") and url not in images:
                images.append(url)

        family = clean_title(raw.get("FAMILIE"))
        category = category_from_categories(raw.get("CATEGORIES"))
        composition = text(raw.get("COMPOSITION"))
        description = text(raw.get("DESCRIPTION"))
        observation = text(raw.get("OBSERVATION"))

        if model not in products:
            specifications = [
                item
                for item in [
                    spec("Composición", composition),
                    spec("Observaciones", observation),
                    spec("Medidas", raw.get("MEASURES")),
                    spec("Familia", family),
                    spec("Categorías", raw.get("CATEGORIES")),
                    spec("Unidades por pack", raw.get("UNITSPACK")),
                    spec("Unidades por caja", raw.get("UNITSBOX")),
                    spec("Pedido mínimo", raw.get("MIN_ORDER_QTY")),
                    spec("Incremento de pedido", raw.get("STEP_ORDER_QTY")),
                    spec("Precio 500-1999 uds", raw.get("PRICE FROM 500 TO 1999 UNITS")),
                    spec("Precio 2000-4999 uds", raw.get("PRICE FROM 2000 TO 4999 UNITS")),
                    spec("Precio >= 5000 uds", raw.get("PRICE >= 5000 UNITS")),
                    spec("Medida caja", raw.get("MEDIDA_CAJA")),
                    spec("Peso caja", raw.get("PESO_CAJA")),
                    spec("Código arancelario", raw.get("TARIC_CODE")),
                    spec("Origen", raw.get("MADE_IN")),
                    spec("Código de impresión", raw.get("PRINTCODE")),
                ]
                if item
            ]

            products[model] = {
                "productNumber": model,
                "name": name,
                "slug": f"{slugify(name)}-{slugify(model)}",
                "description": description,
                "gender": None,
                "material": composition,
                "category": category,
                "subcategory": family,
                "brand": "Roly",
                "images": images[:12],
                "documents": [],
                "specifications": specifications,
                "attributes": {
                    key: value
                    for key, value in {
                        "family": family,
                        "categories": raw.get("CATEGORIES"),
                        "composition": composition,
                        "origin": raw.get("MADE_IN"),
                        "printCode": raw.get("PRINTCODE"),
                    }.items()
                    if text(value)
                },
                "variants": [],
            }
        else:
            existing_images = products[model]["images"]
            for url in images:
                if url not in existing_images and len(existing_images) < 12:
                    existing_images.append(url)

        products[model]["variants"].append(
            {
                "sku": sku,
                "name": " / ".join(part for part in [color, size] if part) or sku,
                "color": color,
                "size": size,
                "priceCents": price,
                "stock": 999,
            }
        )

    normalized = []
    for product in products.values():
        seen_skus = set()
        variants = []
        for variant in product["variants"]:
            if variant["sku"] in seen_skus:
                continue
            seen_skus.add(variant["sku"])
            variants.append(variant)
        product["variants"] = variants
        if variants:
            normalized.append(product)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Normalized {len(normalized)} products with {sum(len(item['variants']) for item in normalized)} variants.")
    if skipped:
        print(f"Skipped {len(skipped)} rows without model, SKU, name or price.")
        print(json.dumps(skipped[:10], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python scripts/normalize-roly-xlsx.py <input.xlsx> <output.json>")
    normalize(Path(sys.argv[1]), Path(sys.argv[2]))
