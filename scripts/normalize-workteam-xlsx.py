from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


def text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def cents(value):
    value = text(value)
    if not value:
        return None
    try:
        return int(round(float(value.replace(",", ".")) * 100))
    except ValueError:
        return None


def slugify(value):
    value = unicodedata.normalize("NFD", value).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[^a-zA-Z0-9]+", "-", value.lower()).strip("-")
    return value or "producte"


def title_case(value):
    value = text(value)
    if not value:
        return None
    return " ".join(part.capitalize() for part in value.split())


def image_urls(row):
    urls = []
    for key in ["IMAGEN1", "IMAGEN2", "IMAGEN3", "IMAGEN4", "IMAGEN5", "IMAGEN6"]:
        value = text(row.get(key))
        if value and value.startswith("http") and value not in urls:
            urls.append(value)
    return urls


def document(type_, title, url):
    url = text(url)
    if not url or not url.startswith("http"):
        return None
    return {"type": type_, "title": title, "url": url}


def normalize(xlsx_path: Path, output_path: Path):
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = [str(value).strip() if value is not None else "" for value in next(rows)]
    indexes = {name: index for index, name in enumerate(header) if name}

    products = OrderedDict()
    skipped = []
    current = {}
    carry_fields = [
        "REFERENCIA",
        "SUBDIVISION",
        "TIPO",
        "DESCRIPCION",
        "COMPOSICION",
        "GRAMAJE",
        "COLOR",
        "PVP1",
        "PVP2",
        "PVP3",
        "IMAGEN1",
        "IMAGEN2",
        "IMAGEN3",
        "IMAGEN4",
        "IMAGEN5",
        "IMAGEN6",
        "MEDIDAS TALLAS",
        "FICHA TECNICA",
    ]

    for row_number, row in enumerate(rows, start=2):
        raw = {name: row[index] if index < len(row) else None for name, index in indexes.items()}
        if not any(text(value) for value in raw.values()):
            continue

        if text(raw.get("REFERENCIA")):
            for field in carry_fields:
                value = text(raw.get(field))
                if value:
                    current[field] = value

        reference = text(current.get("REFERENCIA"))
        size = text(raw.get("TALLA"))
        barcode = text(raw.get("BARCODE"))
        price = cents(current.get("PVP1"))

        if not reference or not size or not price or price <= 0:
            skipped.append({"row": row_number, "reference": reference, "size": size, "price": price})
            continue

        product_type = title_case(current.get("TIPO")) or "Producto"
        name = f"{product_type} {reference}"
        category = title_case(current.get("SUBDIVISION"))
        subcategory = product_type
        color = text(current.get("COLOR"))
        description = text(current.get("DESCRIPCION"))
        material = text(current.get("COMPOSICION"))

        if reference not in products:
            docs = []
            size_guide = document("SIZE_GUIDE", "Guía de tallas", current.get("MEDIDAS TALLAS"))
            technical_sheet = document("TECHNICAL_SHEET", "Ficha técnica", current.get("FICHA TECNICA"))
            for item in [size_guide, technical_sheet]:
                if item:
                    docs.append(item)

            specifications = [
                item
                for item in [
                    {"label": "Composición", "value": material} if material else None,
                    {"label": "Gramaje", "value": text(current.get("GRAMAJE"))} if text(current.get("GRAMAJE")) else None,
                    {"label": "Código Intrastat", "value": text(raw.get("INTRASTAT"))} if text(raw.get("INTRASTAT")) else None,
                ]
                if item
            ]

            products[reference] = {
                "productNumber": reference,
                "name": name,
                "slug": f"{slugify(name)}",
                "description": description,
                "gender": None,
                "material": material,
                "category": category,
                "subcategory": subcategory,
                "brand": "Workteam",
                "images": image_urls(current)[:12],
                "documents": docs,
                "specifications": specifications,
                "attributes": {
                    key: value
                    for key, value in {
                        "subdivision": category,
                        "type": subcategory,
                        "composition": material,
                    }.items()
                    if value
                },
                "variants": [],
            }
        else:
            existing_images = products[reference]["images"]
            for url in image_urls(current):
                if url not in existing_images and len(existing_images) < 12:
                    existing_images.append(url)
            existing_docs = products[reference]["documents"]
            for item in [
                document("SIZE_GUIDE", "Guía de tallas", current.get("MEDIDAS TALLAS")),
                document("TECHNICAL_SHEET", "Ficha técnica", current.get("FICHA TECNICA")),
            ]:
                if item and all(existing["url"] != item["url"] for existing in existing_docs):
                    existing_docs.append(item)

        sku_source = barcode or f"{reference}-{color or 'color'}-{size}"
        products[reference]["variants"].append(
            {
                "sku": sku_source,
                "name": " / ".join(part for part in [color, size] if part) or sku_source,
                "color": color,
                "size": size,
                "priceCents": price,
                "stock": 999,
            }
        )

    normalized = []
    for product in products.values():
        seen_skus = set()
        variants = []
        for variant in product["variants"]:
            if variant["sku"] in seen_skus:
                continue
            seen_skus.add(variant["sku"])
            variants.append(variant)
        product["variants"] = variants
        if variants:
            normalized.append(product)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Normalized {len(normalized)} products with {sum(len(item['variants']) for item in normalized)} variants.")
    if skipped:
        print(f"Skipped {len(skipped)} rows without reference, size or price.")
        print(json.dumps(skipped[:10], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python scripts/normalize-workteam-xlsx.py <input.xlsx> <output.json>")
    normalize(Path(sys.argv[1]), Path(sys.argv[2]))
